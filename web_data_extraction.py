# web_data_extraction.py to extract course data from the given URL and save it to an Excel file

import requests
import openpyxl
from bs4 import BeautifulSoup
from pathlib import Path
import argostranslate.package
import argostranslate.translate

# Configuring variables and parameters
def install_argos_language_pack():
    from_code = "de"
    to_code   = "en"
    argostranslate.package.update_package_index()
    available_packages = argostranslate.package.get_available_packages()
    package = next(
        (p for p in available_packages
         if p.from_code == from_code and p.to_code == to_code),
        None
    )
    if package:
        argostranslate.package.install_from_path(package.download())
        print(f"  Argos language pack installed: {from_code} → {to_code}")
    else:
        print(f"  WARNING: Argos language pack {from_code} → {to_code} not found")

print("Setting up Argos Translate (German → English)")
install_argos_language_pack()

# Static Configurations
URL        = "https://bookstack.cs.ovgu.de/books/msc-digital-engineering-ab-sommer-2026-6hi/page/data-management-for-engineering-applications"
EXCEL_PATH = "Supplementary_Database.xlsx"

# Metadata fields to extract and save in Excel
COLUMNS = [
    "Course Name",
    "LSF Moodle URL of the course",
    "Responsibility",
    "Lecturer",
    "Classes",
    "Applicability in Curriculum",
    "Abbreaviation of the course name",
    "Credit Points",
    "Semester",
    "Term",
    "Duration of the course",
    "Language",
    "Level",
    "Intended Learning Outcomes",
    "Overall Content",
    "Workload",
    "Pre-examination requirements",
    "Type of examination",
    "Teaching method / lecture hours per week (SWS)",
    "Prerequisites according to examination regulations",
    "Recommended prerequisites",
    "Media",
    "Literature",
]

#Creates the Excel file with headers if it does not already exist. If it exists, it will append to it.
def ensure_excel_exists(excel_path, columns):
    path = Path(excel_path)
    if not path.exists():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Module Data"
        ws.append(columns)
        wb.save(str(path))
        print(f"  Created new Excel file with headers: {excel_path}")
    else:
        print(f"  Excel file already exists, appending to: {excel_path}")


# Translates text from German to English using Argos Translate, with error handling and fallback
def translate_text(text_value):
    if not text_value or not text_value.strip():
        return text_value
    try:
        translated = argostranslate.translate.translate(text_value, "de", "en")
        if translated and translated.strip():
            return translated
    except Exception as e:
        print(f"    Translation failed for this item. Reason: {e}")

    print(f"[Translation skipped]: {text_value!r}")
    return text_value

# Fetches the webpage content and returns a BeautifulSoup object for parsing
def fetch(url):
    resp = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=30)
    resp.raise_for_status()
    return BeautifulSoup(resp.text, "lxml")

# Helper function to extract text from a BeautifulSoup tag, handling None values
def text(tag):
    return tag.get_text(separator=" ", strip=True) if tag else ""

# Extracts course data from the BeautifulSoup object and organizes it into a dictionary
def extract(soup, url):
    row = {col: "" for col in COLUMNS}

    body   = soup.find("div", class_="page-content") or soup.find("main") or soup.body
    tables = body.find_all("table")

    h1 = soup.find("h1")
    row["Course Name"] = h1.get_text(strip=True) if h1 else ""
    row["LSF Moodle URL of the course"] = url

    if len(tables) >= 2:
        for tr in tables[1].find_all("tr"):
            cells = tr.find_all(["th", "td"])
            if len(cells) < 2:
                continue
            label = cells[0].get_text(strip=True).rstrip(":").lower()
            val   = text(cells[1])

            if "link" in label:
                anchor = cells[1].find("a")
                if anchor and anchor.get("href"):
                    row["LSF Moodle URL of the course"] = anchor["href"]
                else:
                    row["LSF Moodle URL of the course"] = val
            elif "responsibility" in label:
                row["Responsibility"] = val
            elif "lecturer" in label:
                row["Lecturer"] = val
            elif "classes" in label:
                items = [s.strip() for s in cells[1].get_text("\n", strip=True).splitlines() if s.strip()]
                if len(items) % 2 == 0:
                    paired = [f"{items[i]} {items[i+1]}" for i in range(0, len(items), 2)]
                    row["Classes"] = "; ".join(paired)
                else:
                    row["Classes"] = "; ".join(items)
            elif "applicability" in label or "curriculum" in label:
                row["Applicability in Curriculum"] = val.lstrip("- ").strip()

    if len(tables) >= 3:
        for td in tables[2].find_all(["td", "th"]):
            ps = td.find_all("p")
            if len(ps) >= 2:
                label = ps[0].get_text(strip=True).lower()
                val   = ps[1].get_text(strip=True)
                if   "abbreviation"  in label: row["Abbreaviation of the course name"] = val
                elif "credit point"  in label: row["Credit Points"]                    = val
                elif label == "semester":      row["Semester"]                          = val
                elif label == "term":          row["Term"]                              = val
                elif "duration"      in label: row["Duration of the course"]            = val
                elif "language"      in label: row["Language"]                          = val
                elif "level"         in label: row["Level"]                             = val

    for p in body.find_all("p"):
        t  = p.get_text(separator=" ", strip=True)
        lo = t.lower()

        if lo.startswith("intended learning outcome"):
            parts   = []
            sibling = p.find_next_sibling()
            while sibling:
                if sibling.name == "p":
                    intro = sibling.get_text(strip=True)
                    if intro:
                        parts.append(intro)
                    sibling = sibling.find_next_sibling()
                elif sibling.name == "ul":
                    items = [li.get_text(strip=True) for li in sibling.find_all("li")]
                    parts.extend(items)
                    break
                else:
                    break
            row["Intended Learning Outcomes"] = "; ".join(parts)

        elif lo == "content:":
            ul = p.find_next_sibling("ul")
            if ul:
                row["Overall Content"] = "; ".join(
                    li.get_text(strip=True) for li in ul.find_all("li")
                )

        elif lo.startswith("workload:"):
            row["Workload"] = t.split(":", 1)[1].strip()

    if len(tables) >= 4:
        rows4 = tables[3].find_all("tr")
        if len(rows4) >= 2:
            heads = [text(td).rstrip(":").lower() for td in rows4[0].find_all(["th", "td"])]
            cells = rows4[1].find_all(["th", "td"])
            for h, td in zip(heads, cells):
                if "pre-examination" in h:
                    row["Pre-examination requirements"] = text(td)
                elif "type of examination" in h:
                    row["Type of examination"] = text(td)
                elif "teaching method" in h or "sws" in h or "lecture hours" in h:
                    lis = td.find_all("li")
                    row["Teaching method / lecture hours per week (SWS)"] = (
                        "; ".join(li.get_text(strip=True) for li in lis) if lis else text(td)
                    )

    if len(tables) >= 5:
        rows5 = tables[4].find_all("tr")
        if len(rows5) >= 2:
            heads = [text(td).rstrip(":").lower() for td in rows5[0].find_all(["th", "td"])]
            vals  = [text(td)                     for td in rows5[1].find_all(["th", "td"])]
            for h, v in zip(heads, vals):
                if "prerequisites according" in h:
                    row["Prerequisites according to examination regulations"] = v
                elif "recommended" in h:
                    row["Recommended prerequisites"] = v

    if len(tables) >= 6:
        rows6 = tables[5].find_all("tr")
        if len(rows6) >= 2:
            heads = [text(td).rstrip(":").lower() for td in rows6[0].find_all(["th", "td"])]
            vals  = [text(td)                     for td in rows6[1].find_all(["th", "td"])]
            for h, v in zip(heads, vals):
                if "media"        in h: row["Media"]      = v
                elif "literature" in h: row["Literature"] = v

    return row

# Appends the extracted and translated data as a new row in the Excel file
def append_to_excel(row, excel_path):
    path = Path(excel_path)
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    ws.append([row.get(h, "") for h in headers])
    wb.save(str(path))
    print(f"  Data saved — {ws.max_row - 1} row(s) in file.")


# main execution flow

ensure_excel_exists(EXCEL_PATH, COLUMNS)
soup = fetch(URL)
row  = extract(soup, URL)
# Translate
print("\nStarting translation of all fields to English...")
for field in COLUMNS:
    original = row.get(field, "")
    if original:
        translated = translate_text(original)
        if translated != original:
            print(f"  Translated '{field}':\n    Original (DE): {original!r}\n    Translated (EN): {translated!r}")
        row[field] = translated

#Printing extracted fields
print("\nHere are the extracted fields:")
for col, val in row.items():
    print(f"  {col}: {val!r}")

# Append the data to Excel
append_to_excel(row, EXCEL_PATH)
print(f"\nEntry added to: {EXCEL_PATH}")
print(f"Course Name: {row.get('Course Name', '')}")